"""Merge old and Stage 1 PRS CSV logs into one Excel workbook.

Usage:
    python CSV_merge.py [CSV_FOLDER] [-o OUTPUT.xlsx]

Dependencies:
    python -m pip install pandas openpyxl
"""

from __future__ import annotations

import argparse
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as error:
    raise SystemExit(
        f"Missing Python package: {error.name}. "
        "Install dependencies with: python -m pip install pandas openpyxl"
    ) from error


MAX_COLUMN_WIDTH = 100
FONT_NAME = "Calibri"
FONT_SIZE = 11
HEADER_FONT_SIZE = 12


def get_unique_sheet_name(name: str, used_names: set[str]) -> str:
    sheet_name = Path(name).stem
    for char in "\\/*?:[]":
        sheet_name = sheet_name.replace(char, "_")
    sheet_name = sheet_name[:31] or "Sheet"

    original_name = sheet_name
    counter = 1
    while sheet_name in used_names:
        suffix = f"_{counter}"
        sheet_name = original_name[: 31 - len(suffix)] + suffix
        counter += 1
    used_names.add(sheet_name)
    return sheet_name


def format_worksheet(worksheet) -> None:
    for row in worksheet.iter_rows():
        for cell in row:
            cell.font = Font(name=FONT_NAME, size=FONT_SIZE)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for cell in worksheet[1]:
        cell.font = Font(name=FONT_NAME, size=HEADER_FONT_SIZE, bold=True)

    for column_cells in worksheet.columns:
        max_length = max(
            (max(len(line) for line in str(cell.value).split("\n"))
             for cell in column_cells if cell.value is not None),
            default=0,
        )
        column_letter = get_column_letter(column_cells[0].column)
        worksheet.column_dimensions[column_letter].width = min(
            max_length + 3, MAX_COLUMN_WIDTH
        )

    worksheet.row_dimensions[1].height = 22
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def read_csv_files(input_folder: Path) -> dict[str, pd.DataFrame]:
    frames: dict[str, pd.DataFrame] = {}
    for csv_file in sorted(input_folder.glob("*.csv")):
        try:
            frame = pd.read_csv(csv_file, encoding="utf-8-sig")
            frame.insert(0, "source_file", csv_file.name)
            frames[csv_file.name] = frame
            print(f"Loaded {csv_file.name}: {len(frame)} rows")
        except Exception as error:
            print(f"Skipped {csv_file.name}: {error}")
    return frames


def concatenate_schema(
    frames: dict[str, pd.DataFrame], required_columns: set[str]
) -> pd.DataFrame:
    matching = [
        frame.copy()
        for frame in frames.values()
        if required_columns.issubset(frame.columns)
    ]
    return pd.concat(matching, ignore_index=True) if matching else pd.DataFrame()


def numeric_id(frame: pd.DataFrame, column: str) -> pd.Series:
    if column not in frame:
        return pd.Series(pd.NA, index=frame.index, dtype="Int64")
    return pd.to_numeric(frame[column], errors="coerce").astype("Int64")


def numeric_column(
    frame: pd.DataFrame, column: str, default: float = 0.0
) -> pd.Series:
    if column not in frame:
        return pd.Series(default, index=frame.index, dtype="float64")
    return pd.to_numeric(frame[column], errors="coerce")


def prefix_except(
    frame: pd.DataFrame, prefix: str, retained: set[str]
) -> pd.DataFrame:
    return frame.rename(
        columns={column: f"{prefix}{column}" for column in frame if column not in retained}
    )


def build_attempt_summary(frames: dict[str, pd.DataFrame]) -> pd.DataFrame:
    acquisition = concatenate_schema(frames, {"node", "failure_reason"})
    if acquisition.empty:
        return acquisition

    acquisition["attempt_id"] = numeric_id(acquisition, "attempt_id")
    acquisition["poll_frame_id"] = numeric_id(acquisition, "poll_frame_id")

    initiator = acquisition[acquisition["node"] == "initiator"].copy()
    responder = acquisition[acquisition["node"] == "responder"].copy()
    initiator["merge_attempt_id"] = initiator["attempt_id"].fillna(
        initiator["poll_frame_id"]
    )
    responder["merge_attempt_id"] = responder["attempt_id"].fillna(
        responder["poll_frame_id"]
    )

    # A responder CRC-failure ID came from a failed payload and is not trusted.
    responder["merge_id_trusted"] = (
        pd.to_numeric(responder.get("frame_id_valid", 0), errors="coerce") == 1
    ) | responder["attempt_id"].notna()

    for frame in (initiator, responder):
        frame.dropna(subset=["merge_attempt_id"], inplace=True)
        frame["attempt_occurrence"] = frame.groupby("merge_attempt_id").cumcount()

    retained = {"merge_attempt_id", "attempt_occurrence"}
    initiator = prefix_except(initiator, "initiator_", retained)
    responder = prefix_except(responder, "responder_", retained)
    summary = initiator.merge(
        responder,
        on=["merge_attempt_id", "attempt_occurrence"],
        how="outer",
        validate="one_to_one",
    )

    def classify(row: pd.Series) -> str:
        responder_reason = row.get("responder_failure_reason")
        initiator_reason = row.get("initiator_failure_reason")
        if responder_reason == "PAYLOAD_CRC":
            return "POLL_PAYLOAD_CRC"
        if responder_reason == "NONE" and initiator_reason == "NONE":
            return "SUCCESS"
        if initiator_reason == "PAYLOAD_CRC":
            return "RESPONSE_PAYLOAD_CRC"
        if initiator_reason == "NO_PREAMBLE":
            return "NO_RESPONSE"
        return "UNMATCHED_OR_OTHER"

    summary.insert(2, "end_to_end_result", summary.apply(classify, axis=1))
    return summary.sort_values(["merge_attempt_id", "attempt_occurrence"])


def build_paired_measurements(frames: dict[str, pd.DataFrame]) -> pd.DataFrame:
    measurements = concatenate_schema(
        frames, {"direction", "poll_frame_id", "fine_delay_samples"}
    )
    if measurements.empty:
        return measurements

    measurements["poll_frame_id"] = numeric_id(measurements, "poll_frame_id")
    valid = pd.to_numeric(
        measurements.get("frame_id_valid", 0), errors="coerce"
    ) == 1
    measurements = measurements[valid & measurements["poll_frame_id"].notna()].copy()

    poll = measurements[measurements["direction"] == "poll_rx"].copy()
    response = measurements[measurements["direction"] == "response_rx"].copy()
    for frame in (poll, response):
        frame["measurement_occurrence"] = frame.groupby("poll_frame_id").cumcount()

    retained = {"poll_frame_id", "measurement_occurrence"}
    poll = prefix_except(poll, "poll_", retained)
    response = prefix_except(response, "response_", retained)
    paired = poll.merge(
        response,
        on=["poll_frame_id", "measurement_occurrence"],
        how="outer",
        validate="one_to_one",
    )

    poll_contribution = numeric_column(
        paired, "poll_phase_range_contribution_m"
    ).fillna(0.0)
    response_contribution = numeric_column(
        paired, "response_phase_range_contribution_m"
    ).fillna(0.0)
    integer_range = numeric_column(
        paired, "response_integer_range_m", float("nan")
    )
    paired.insert(
        2,
        "diagnostic_phase_corrected_range_m",
        integer_range + poll_contribution + response_contribution,
    )
    paired.insert(
        3,
        "has_both_directions",
        paired.get("poll_direction").notna() & paired.get("response_direction").notna(),
    )
    return paired.sort_values(["poll_frame_id", "measurement_occurrence"])


def build_column_guide() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "column": "prs_cp_cfo_coherence",
                "range": "0 to 1",
                "meaning": (
                    "Normalized correlation between every OFDM cyclic prefix "
                    "and its matching useful-symbol tail, combined over all PRS symbols."
                ),
                "interpretation": (
                    "Near 1 means a strong, consistently phased CP match. Low values can "
                    "indicate noise, incorrect timing, interference, or a channel exceeding the CP."
                ),
            },
            {
                "column": "channel_coherence",
                "range": "0 to 1",
                "meaning": (
                    "Normalized adjacent-symbol correlation of H_m[k]=Y_m[k]/X_m[k] "
                    "over all PRS bins and symbol pairs."
                ),
                "interpretation": (
                    "Near 1 means the channel is stable across the 16 PRS symbols apart "
                    "from common CFO phase rotation. It is a stability metric, not SNR or dBm."
                ),
            },
            {
                "column": "residual_cfo_hz",
                "range": "Hz",
                "meaning": "PRS channel CFO minus the CFO reference used for phase unwrapping.",
                "interpretation": "Large magnitude suggests disagreement between CFO estimators.",
            },
            {
                "column": "phase_range_contribution_m",
                "range": "metres",
                "meaning": "One-direction diagnostic contribution c*fine_delay/2.",
                "interpretation": "Requires the matched opposite direction and RF calibration.",
            },
            {
                "column": "diagnostic_phase_corrected_range_m",
                "range": "metres",
                "meaning": (
                    "Integer response range plus matched poll and response phase contributions."
                ),
                "interpretation": "Diagnostic only; sign and RF group delay are not calibrated.",
            },
            {
                "column": "responder_merge_id_trusted",
                "range": "TRUE/FALSE",
                "meaning": "Whether the responder attempt ID came from a valid payload or gate.",
                "interpretation": "FALSE IDs decoded from CRC-failed payloads may be incorrect.",
            },
        ]
    )


def write_workbook(
    output_file: Path,
    frames: dict[str, pd.DataFrame],
    attempt_summary: pd.DataFrame,
    paired_measurements: pd.DataFrame,
) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    used_sheet_names: set[str] = set()
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        build_column_guide().to_excel(writer, sheet_name="column_guide", index=False)
        used_sheet_names.add("column_guide")
        if not attempt_summary.empty:
            attempt_summary.to_excel(writer, sheet_name="attempt_summary", index=False)
            used_sheet_names.add("attempt_summary")
        if not paired_measurements.empty:
            paired_measurements.to_excel(
                writer, sheet_name="paired_prs_measurements", index=False
            )
            used_sheet_names.add("paired_prs_measurements")
        for filename, frame in frames.items():
            frame.to_excel(
                writer,
                sheet_name=get_unique_sheet_name(filename, used_sheet_names),
                index=False,
            )

    workbook = load_workbook(output_file)
    for worksheet in workbook.worksheets:
        format_worksheet(worksheet)
    workbook.save(output_file)


def parse_args() -> argparse.Namespace:
    script_folder = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description="Merge PRS CSV logs into a formatted Excel workbook."
    )
    parser.add_argument(
        "input_folder",
        nargs="?",
        type=Path,
        default=script_folder,
        help="Folder containing CSV logs (default: this script's folder).",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="Output .xlsx path (default: INPUT_FOLDER/Merged.xlsx).",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_folder = args.input_folder.resolve()
    output_file = (args.output or input_folder / "Merged.xlsx").resolve()
    if not input_folder.is_dir():
        print(f"Input folder does not exist: {input_folder}")
        return 1

    frames = read_csv_files(input_folder)
    if not frames:
        print(f"No readable CSV files found in {input_folder}")
        return 1

    attempt_summary = build_attempt_summary(frames)
    paired_measurements = build_paired_measurements(frames)
    write_workbook(output_file, frames, attempt_summary, paired_measurements)
    print(f"Wrote {output_file}")
    print(f"Raw CSV sheets: {len(frames)}")
    print(f"Attempt summary rows: {len(attempt_summary)}")
    print(f"Paired PRS rows: {len(paired_measurements)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
